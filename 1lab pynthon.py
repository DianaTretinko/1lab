import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\Diana Tr\Desktop\testi\test.xlsx')
ws = wb.active
max_row = ws.max_row

count_above_3000 = 0

for row in range(2, max_row + 1):
    hours = ws['B' + str(row)].value
    rate = ws['C' + str(row)].value

    try:
       
        if hours is not None and rate is not None and isinstance(hours, (int, float)) and isinstance(rate, (int, float)):
            salary = hours * rate
            if salary > 3000:
                count_above_3000 += 1
        else:
            print(f"Invalid data in row {row}. Skipping.")
    except TypeError:
        print(f"Invalid data type in row {row}. Skipping.")

print("People with a salary greater than 3000 EUR:", count_above_3000)